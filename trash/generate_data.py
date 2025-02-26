import pandas as pd
import numpy as np
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

def generate_fitness_dataset(num_records=200):
    # Possible values for each column
    meal_plans = ['Weight Loss', 'Muscle Gain', 'Maintain Health']
    experience_levels = ['Beginner', 'Moderate', 'Expert']
    bmi_categories = ['Underweight', 'Normal', 'Overweight', 'Obese']
    activity_levels = ['Not Active', 'Quite Active', 'Active']
    equipment_preferences = ['No Equipment', 'With Equipment', 'Both Works Fine']
    dietary_restrictions = [
        'None', 'Vegan', 'Gluten-Free', 'Low-Sodium', 
        'Dairy-Free', 'Vegetarian', 'Pescatarian'
    ]

    # Generate dataset
    data = []
    for i in range(num_records):
        # Strategic mapping of attributes
        if i < 50:  # First 50: Weight Loss focus
            meal_plan = 'Weight Loss'
            bmi = random.choices(['Overweight', 'Obese'], weights=[0.6, 0.4])[0]
            experience = random.choices(experience_levels, weights=[0.5, 0.3, 0.2])[0]
        elif i < 100:  # Next 50: Muscle Gain focus
            meal_plan = 'Muscle Gain'
            bmi = random.choices(['Normal', 'Underweight'], weights=[0.7, 0.3])[0]
            experience = random.choices(experience_levels, weights=[0.3, 0.4, 0.3])[0]
        else:  # Remaining: Maintain Health
            meal_plan = 'Maintain Health'
            bmi = 'Normal'
            experience = random.choices(experience_levels, weights=[0.4, 0.4, 0.2])[0]

        # Strategic training duration mapping
        if experience == 'Beginner':
            training_duration = random.choices([20, 30], weights=[0.6, 0.4])[0]
        elif experience == 'Moderate':
            training_duration = random.choices([30, 45], weights=[0.5, 0.5])[0]
        else:  # Expert
            training_duration = random.choices([45, 60], weights=[0.4, 0.6])[0]

        # Create record
        record = {
            'No': i + 1,
            'Meal & Workout Plan': meal_plan,
            'Experience Level': experience,
            'BMI Baseline': bmi,
            'Activity Level': random.choice(activity_levels),
            'Training Duration (min/day)': training_duration,
            'Equipment Preference': random.choice(equipment_preferences),
            'Dietary Restrictions/Allergies': random.choice(dietary_restrictions)
        }
        data.append(record)

    # Create DataFrame
    df = pd.DataFrame(data)
    return df

def create_formatted_excel(df):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Fitness Profiles"

    # Create additional worksheets
    ws2 = wb.create_sheet("Summary")
    ws3 = wb.create_sheet("Distribution")

    # Formatting for main dataset worksheet
    headers = list(df.columns)
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for r_idx, row in enumerate(df.values.tolist(), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].auto_size = True

    # Create Summary Worksheet
    summary_data = [
        ["Dataset Metrics"],
        ["Total Records", len(df)],
        ["Columns", len(df.columns)]
    ]

    for r, row in enumerate(summary_data, 1):
        for c, value in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=value)
            if r == 1:
                cell.font = Font(bold=True, size=14)

    # Distribution Worksheet
    distribution_sheets = {
        "Meal & Workout Plan": df['Meal & Workout Plan'].value_counts(),
        "Experience Level": df['Experience Level'].value_counts(),
        "BMI Baseline": df['BMI Baseline'].value_counts(),
        "Activity Level": df['Activity Level'].value_counts(),
        "Equipment Preference": df['Equipment Preference'].value_counts(),
        "Dietary Restrictions": df['Dietary Restrictions/Allergies'].value_counts()
    }

    row = 1
    for sheet_name, series in distribution_sheets.items():
        ws3.cell(row=row, column=1, value=sheet_name).font = Font(bold=True)
        row += 1
        
        for index, value in series.items():
            ws3.cell(row=row, column=1, value=str(index))
            ws3.cell(row=row, column=2, value=value)
            row += 1
        
        row += 2  # Add a blank row between distributions

    # Save the workbook
    wb.save('fitness_profiles_dataset_detailed.xlsx')
    print("Excel file 'fitness_profiles_dataset_detailed.xlsx' created successfully!")

# Generate and save the dataset
fitness_data = generate_fitness_dataset(1000)1
create_formatted_excel(fitness_data)

# Optional: Display some summary information
print("\nDataset Summary:")
print(fitness_data['Meal & Workout Plan'].value_counts())
print("\nBMI Distribution:")
print(fitness_data['BMI Baseline'].value_counts())
